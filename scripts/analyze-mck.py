import openpyxl
import os
import sys
import csv
import tempfile
from io import BytesIO

def clean_value(val):
    if val is None: return ""
    return str(val).strip()

def clean_number(val):
    if val is None or val == "": return 0.0
    try:
        # Handle Indonesian format if string
        if isinstance(val, str):
            val = val.replace('.', '').replace(',', '.')
        return float(val)
    except:
        return 0.0

def extract_rab_mck(file_path, sheet_name="RAB", start_row=77):
    if not os.path.exists(file_path):
        print(f"Error: File {file_path} not found.", file=sys.stderr)
        sys.exit(1)
        
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        # Try finding a sheet that contains 'RAB'
        found = False
        for name in wb.sheetnames:
            if 'RAB' in name.upper():
                sheet_name = name
                found = True
                break
        if not found:
            sheet_name = wb.sheetnames[0]
        
    sheet = wb[sheet_name]
    
    # Try to find "PEKERJAAN" in column 2, rows 1-75
    pekerjaan_name = ""
    for r in range(1, 76):
        cell_val = str(sheet.cell(row=r, column=2).value or "").strip().lower()
        if "pekerjaan" in cell_val:
            val = sheet.cell(row=r, column=4).value
            if val:
                pekerjaan_name = str(val).strip().lstrip(':').strip()
                break
    
    # Try to find header at row 76
    header_row = 76
    headers = {}
    for cell in sheet[header_row]:
        val = str(cell.value).strip().lower() if cell.value else ""
        if val:
            headers[val] = cell.column
            
    # Define targets aliases
    target_fields = {
        "item": ["item pekerjaan", "uraian pekerjaan", "uraian", "pekerjaan"],
        "satuan": ["satuan", "unit"],
        "volume": ["volume", "vol"],
        "harga": ["harga satuan", "harga", "unit price"]
    }
    
    col_map = {}
    for field, aliases in target_fields.items():
        for alias in aliases:
            if alias in headers:
                col_map[field] = headers[alias]
                break
    
    # Fallback to defaults
    if "item" not in col_map: col_map["item"] = 2
    if "satuan" not in col_map: col_map["satuan"] = 3
    if "volume" not in col_map: col_map["volume"] = 4
    if "harga" not in col_map: col_map["harga"] = 5
    
    data = []
    grand_total = 0
    
    # Start from row 77
    for row_idx in range(start_row, sheet.max_row + 1):
        if sheet.row_dimensions[row_idx].hidden:
            continue
            
        item = sheet.cell(row=row_idx, column=col_map["item"]).value
        satuan = sheet.cell(row=row_idx, column=col_map["satuan"]).value
        volume = sheet.cell(row=row_idx, column=col_map["volume"]).value
        harga = sheet.cell(row=row_idx, column=col_map["harga"]).value
        
        if item:
            item_str = str(item).strip()
            if item_str.lower() == "jumlah" or "total" in item_str.lower():
                continue
                
            # Normalize units
            item_str = item_str.replace("m²", "m2").replace("m³", "m3")
            satuan_str = (str(satuan).strip() if satuan else "").replace("m²", "m2").replace("m³", "m3")
            
            vol_num = clean_number(volume)
            harga_num = clean_number(harga)
            total_num = vol_num * harga_num
            # grand_total += total_num (We'll use H78 for the official total as per user)
            
            is_header = (harga is None or harga == "")
                
            data.append({
                "item": item_str,
                "satuan": satuan_str,
                "volume": vol_num,
                "harga": harga_num,
                "pajak": "11" if not is_header else "",
                "keterangan": "",
                "kunci": "FALSE" if not is_header else "TRUE",
                "total": total_num,
                "type": "header" if is_header else "item"
            })

    # Get document total from H78 (Column 8, Row 78) as requested by user
    doc_total_val = sheet.cell(row=78, column=8).value
    document_total = clean_number(doc_total_val)
            
    # Final CSV rows
    final_rows = []
    # Project Name row
    if pekerjaan_name:
        # Col order: Item, Satuan, Volume, Harga, Pajak, Keterangan, Kunci, Total, Type
        final_rows.append([pekerjaan_name, "", "", "", "", "", "TRUE", "0", "header"])
        
    for d in data:
        final_rows.append([
            d["item"], 
            d["satuan"], 
            str(d["volume"]) if d["volume"] else "", 
            str(d["harga"]) if d["harga"] else "", 
            d["pajak"], 
            d["keterangan"], 
            d["kunci"],
            str(d["total"]),
            d["type"]
        ])
        
    # Metadata row for validation
    final_rows.append(["METADATA_TOTAL", "GRAND_TOTAL", "", "", "", "", "", str(document_total), "summary"])
    
    return final_rows

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python analyze_mck.py <input_file>")
        sys.exit(1)
        
    input_file = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    try:
        rows = extract_rab_mck(input_file)
        
        if output_path:
            # Ensure directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            path = output_path
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Item", "Satuan", "Volume", "Harga", "Pajak", "Keterangan", "Kunci", "Total", "Type"])
                writer.writerows(rows)
        else:
            # Create temp CSV file
            fd, path = tempfile.mkstemp(suffix='.csv', prefix='rab_analysis_')
            with os.fdopen(fd, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                # Match the new order: Item, Satuan, Volume, Harga, Pajak, Keterangan, Kunci, Total, Type
                writer.writerow(["Item", "Satuan", "Volume", "Harga", "Pajak", "Keterangan", "Kunci", "Total", "Type"])
                writer.writerows(rows)
            
        print(path)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
